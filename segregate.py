import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

file_path = r"filepath--"

# === Create extra sheets ===
df = pd.read_excel(file_path, engine="openpyxl")

# Sheets by Agent Code (only if >= 40 rows)
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    for value in df["Agent Cd"].unique():
        subset = df[df["Agent Cd"] == value]
        if len(subset) >= 40:
            subset.to_excel(writer, sheet_name=str(value), index=False)
print("Additional sheets created in the same Excel file with more than 40")

# High valued policies
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    high_value_subset = df[df["Instalment"] > 20000]
    high_value_subset.to_excel(writer, sheet_name="HIGH", index=False)
print("High Valued Policies are classified into an extra sheet")

# === Highlight DOC before cutoff date ===
cutoff_date = datetime(2023, 4, 30).date()
wb = load_workbook(file_path)
highlight_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Find DOC column index
    doc_col_index = None
    for cell in ws[1]:
        if str(cell.value).strip().lower() == "doc":
            doc_col_index = cell.column
            break
    if doc_col_index is None:
        continue

    for row in ws.iter_rows(min_row=2):
        cell = row[doc_col_index - 1]
        cell_value = cell.value

        doc_date = None
        if isinstance(cell_value, datetime):
            doc_date = cell_value.date()
        elif isinstance(cell_value, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    doc_date = datetime.strptime(cell_value.strip(), fmt).date()
                    break
                except ValueError:
                    continue

        # Apply highlight if before cutoff
        if doc_date and doc_date < cutoff_date:
            for c in row:
                c.fill = highlight_fill

print("DOC before the financial year are highlighted")

sensitive_columns = ['Customer Name', 'Mobile No']

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in ws[1]:
        if col.value in sensitive_columns:
            ws.column_dimensions[get_column_letter(col.column)].hidden = True
            
wb.save(file_path)
print("Sensitive Information Hidden")
